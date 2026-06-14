import pandas as pd
from typing import Dict, List, Literal
from pydantic import BaseModel
from langchain_core.messages import AIMessage, HumanMessage, SystemMessage
from langchain_anthropic import ChatAnthropic
from enum import Enum
from datetime import datetime, date


class SheetClassification(BaseModel):
    classifications: Dict[str, Literal[
        "financial_statements",
        "bureau_data", 
        "banking",
        "emi_table",
        "debtor_creditor",
        "scoring",
        "irrelevant"
    ]]

class DataIngestor:

    # Column name patterns that identify person (director/promoter) columns.
    # Assumes entity names appear in structured columns — not freeform cell text.
    PERSON_COLUMN_PATTERNS = [
        # Patterns redacted from public version.
        # These encode domain-specific column name patterns from Indian MSME CAMs.
    ]

    # Column name patterns that identify related/group entity columns.
    # Deliberately specific to avoid false positives (e.g. "associate director").
    ENTITY_COLUMN_PATTERNS = [
        # Patterns redacted from public version.
        # These encode domain-specific column name patterns from Indian MSME CAMs.
    ]

    # Position values that identify a row as containing a person's name.
    # Used in position-column scan: if a "Position" column contains one of these
    # values, the "Name" column in the same row is extracted as a person.
    PERSON_POSITION_VALUES = [
        # Patterns redacted from public version.
        # These encode domain-specific role values from Indian MSME CAMs.
    ]

    def __init__(self):
        self.llm_0 = ChatAnthropic(model="claude-haiku-4-5-20251001", verbose=True)
        self.llm_0_with_structured_output = self.llm_0.with_structured_output(SheetClassification)
    
    def classify_sheets_with_llm(self, sheet_samples) -> Dict:
        # SYSTEM PROMPT REDACTED
        # The sheet classification prompt instructs the model to assign each sheet to
        # one of the seven categories defined in SheetClassification, based on column
        # names and preview data from the uploaded CAM workbook.
        # In production this is loaded from a config file not included in this repo.
        system_message = ""

        user_message = f"""Here are the sheet samples:\n {sheet_samples}\n"""
        messages = [SystemMessage(content=system_message), HumanMessage(content=user_message)]
        response = self.llm_0_with_structured_output.invoke(messages)

        return response.classifications

    def build_entity_mapping(self, sheets: Dict[str, pd.DataFrame]) -> Dict[str, str]:
        """
        Scan all sheets for entity names in structured columns.
        Returns mapping of real_name -> placeholder (e.g. "Rajesh Sharma" -> "PERSON_1").

        Four scan patterns:
        1. Column-header scan: column name matches PERSON_COLUMN_PATTERNS or
           ENTITY_COLUMN_PATTERNS → extract all non-empty values.
        2. Key-value row scan: first column is a parameter label (e.g. "Director 1"),
           second column is the value → extract if label matches patterns.
        3. Position-column scan: sheet has both "Name" and "Position"/"Designation"
           columns → extract Name values where Position matches a known person role.
        4. Mid-sheet header scan: detects sub-table headers embedded mid-sheet
           (e.g. directors table at row 23 within a larger sheet). Reconstructs
           a header=None view and applies position-column logic to each sub-table.

        Prerequisites:
        - Entity names must appear in structured columns with recognisable column names.
        - Freeform cell text containing entity names is not substituted.
          Document this assumption in COMPLIANCE_GAPS.md.
        """
        persons = []
        entities = []

        for sheet_name, df in sheets.items():
            # --- Column-header scan ---
            # Column names match PERSON or ENTITY patterns → extract all values.
            for col in df.columns:
                col_lower = str(col).lower()

                if any(p in col_lower for p in self.PERSON_COLUMN_PATTERNS):
                    for val in df[col].dropna():
                        val_str = str(val).strip()
                        if val_str and val_str.lower() not in ("nan", "none", ""):
                            persons.append(val_str)

                elif any(p in col_lower for p in self.ENTITY_COLUMN_PATTERNS):
                    for val in df[col].dropna():
                        val_str = str(val).strip()
                        if val_str and val_str.lower() not in ("nan", "none", ""):
                            entities.append(val_str)

            # --- Key-value row scan ---
            # Handles scoring sheets structured as parameter/value pairs:
            # e.g. row = ("Director 1", "Rajesh Sharma")
            # Checks first column as key, second column as value.
            if len(df.columns) >= 2:
                for _, row in df.iterrows():
                    key = str(row.iloc[0]).lower() if pd.notna(row.iloc[0]) else ""
                    if len(row) < 2 or pd.isna(row.iloc[1]):
                        continue
                    val_str = str(row.iloc[1]).strip()
                    if not val_str or val_str.lower() in ("nan", "none", ""):
                        continue

                    if any(p in key for p in self.PERSON_COLUMN_PATTERNS):
                        persons.append(val_str)
                    elif any(p in key for p in self.ENTITY_COLUMN_PATTERNS):
                        entities.append(val_str)

            # --- Position-column scan ---
            # Handles tables where "Name" and "Position" are separate columns.
            # Structure seen in real CAMs: directors table with columns
            # ["Name", "Position", "Age", ...] where Position = "Director" / "Partner" etc.
            # If a "Position" column exists alongside a "Name" column, extract Name values
            # only for rows where Position matches a known person role.
            col_names_lower = [str(c).lower() for c in df.columns]
            name_col = next(
                (c for c in df.columns if str(c).lower().strip() == "name"), None
            )
            position_col = next(
                (c for c in df.columns if str(c).lower().strip() in ("position", "designation")), None
            )
            if name_col is not None and position_col is not None:
                for _, row in df.iterrows():
                    pos_val = str(row[position_col]).lower().strip() if pd.notna(row[position_col]) else ""
                    name_val = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
                    if (
                        name_val and name_val.lower() not in ("nan", "none", "name", "")
                        and any(p in pos_val for p in self.PERSON_POSITION_VALUES)
                    ):
                        persons.append(name_val)

            # --- Option C dedup: bare "Name" column without a Position/Designation column ---
            # Covers director names repeated in Consumer CIBIL bureau sections where the
            # table has only a "Name" column with no role column alongside it.
            # Only adds a name if it already appears in the persons list — prevents pulling
            # in counterparty, bank, or lender names from other Name columns.
            elif name_col is not None and position_col is None:
                existing_persons = set(persons)
                for val in df[name_col].dropna():
                    val_str = str(val).strip()
                    if (
                        val_str
                        and val_str.lower() not in ("nan", "none", "name", "")
                        and val_str in existing_persons
                    ):
                        persons.append(val_str)

        # --- Mid-sheet header scan ---
        # Handles sheets with multiple stacked tables, each with its own header row
        # embedded mid-sheet (e.g. a directors table starting at row 23 within a
        # larger Borrower sheet). pd.read_excel() treats only row 1 as headers, so
        # embedded sub-table headers appear as data rows — the column-header and
        # position-column scans above never fire on them.
        #
        # Strategy: reconstruct a header=None view from the already-loaded df by
        # prepending the column headers as row 0. Scan every row for a cell containing
        # "name" alongside "designation" or "position" in the same row. When found,
        # treat it as a sub-table header and extract Name values from rows below where
        # the role column contains a known PERSON_POSITION_VALUE.
        for sheet_name, df in sheets.items():
            raw_rows = [list(df.columns)] + df.values.tolist()

            for row_idx, row in enumerate(raw_rows):
                row_strs = [str(c).lower().strip() if c is not None else "" for c in row]

                has_name = any(c == "name" for c in row_strs)
                has_role_col = any(c in ("designation", "position") for c in row_strs)
                if not (has_name and has_role_col):
                    continue

                name_idx = next(i for i, c in enumerate(row_strs) if c == "name")
                role_idx = next(
                    i for i, c in enumerate(row_strs)
                    if c in ("designation", "position")
                )

                for data_row in raw_rows[row_idx + 1:]:
                    if all(
                        (c is None or str(c).strip() in ("", "nan", "None"))
                        for c in data_row
                    ):
                        break
                    if name_idx >= len(data_row) or role_idx >= len(data_row):
                        continue
                    name_val = str(data_row[name_idx]).strip() if data_row[name_idx] is not None else ""
                    role_val = str(data_row[role_idx]).lower().strip() if data_row[role_idx] is not None else ""
                    if (
                        name_val
                        and name_val.lower() not in ("nan", "none", "name", "")
                        and any(p in role_val for p in self.PERSON_POSITION_VALUES)
                    ):
                        persons.append(name_val)

        # Build mapping, deduplicating across both lists.
        seen = set()
        mapping = {}

        p_count = 1
        for name in persons:
            if name not in seen:
                seen.add(name)
                mapping[name] = f"PERSON_{p_count}"
                p_count += 1

        e_count = 1
        for name in entities:
            if name not in seen:
                seen.add(name)
                mapping[name] = f"ENTITY_{e_count}"
                e_count += 1

        return mapping

    def _col_to_letter(self, col_idx: int) -> str:
        """Convert zero-based column index to Excel column letter (0→A, 25→Z, 26→AA)."""
        result = ""
        col_idx += 1  # 1-based
        while col_idx:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def read_raw_sheets(self, excel_path: str) -> Dict[str, pd.DataFrame]:
        """
        Read all sheets from an Excel file into raw DataFrames.

        Uses openpyxl directly to handle merged cells correctly —
        pandas read_excel collapses merged regions to NaN in non-top-left cells,
        which breaks PII detection for labels and values in merged layouts.

        Merge-aware: only the top-left cell of each merged region retains its value.
        All other cells in the merge are set to None so they are skipped by
        detect_pii_raw() — preventing duplicate hits for the same merged field.

        No substitution, no classification — raw data only.
        Called on file upload before any other processing.
        """
        import openpyxl

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        raw_sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Build merge map: top-left cell of each merged region keeps its value;
            # all other cells in the region are suppressed to None.
            # This prevents duplicate PII hits for labels/values spanning merged columns.
            merge_map = {}
            for merge_range in ws.merged_cells.ranges:
                top_left_value = ws.cell(merge_range.min_row, merge_range.min_col).value
                for row in range(merge_range.min_row, merge_range.max_row + 1):
                    for col in range(merge_range.min_col, merge_range.max_col + 1):
                        if row == merge_range.min_row and col == merge_range.min_col:
                            merge_map[(row, col)] = top_left_value  # top-left: preserve
                        else:
                            merge_map[(row, col)] = None  # rest of merge: suppress

            data = []
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    if (cell.row, cell.column) in merge_map:
                        row_data.append(merge_map[(cell.row, cell.column)])
                    else:
                        row_data.append(cell.value)
                data.append(row_data)

            df = pd.DataFrame(data)
            df = df.dropna(how="all").dropna(axis=1, how="all")
            raw_sheets[sheet_name] = df

        return raw_sheets

    def detect_pii_raw(self, raw_sheets: Dict[str, pd.DataFrame]) -> List[str]:
        """
        Scan raw DataFrames for PII before any LLM call.
        Returns list of cell references with PII type, e.g.:
            "Scoring!A3 — possible CIN (U15141MP2009PTC022797)"
            "Scoring!B3 — possible PAN (AACCI2746N)"
            "Scoring!A2 — company name field ('Name of Borrower')"

        Called on file upload, before entity mapping, substitution, or any LLM call.
        If any hits are returned, the upload handler shows them to the user and
        requires confirmation before proceeding to analysis.

        Reads with header=None so row indices map directly to Excel rows (row 0 = Excel row 1).
        """
        import re

        CIN_PATTERN = re.compile(r'\b[UL]\d{5}[A-Z]{2}\d{4}[A-Z]{3}\d{6}\b')
        PAN_PATTERN = re.compile(r'\b[A-Z]{5}\d{4}[A-Z]\b')
        COMPANY_NAME_TRIGGER = "name"
        COMPANY_NAME_ENTITY_TOKENS = [
            "borrower", "company", "entity", "business",
            "firm", "applicant", "proprietor", "unit",
            "organisation", "organization", "customer", "client"
        ]

        hits = []

        for sheet_name, df in raw_sheets.items():
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    if not isinstance(value, str) or not value.strip():
                        continue

                    val_upper = value.strip().upper()
                    cell_ref = f"{sheet_name}!{self._col_to_letter(col_idx)}{row_idx + 1}"

                    if CIN_PATTERN.search(val_upper):
                        hits.append(
                            f"{cell_ref} — possible CIN ({value.strip()})"
                        )

                    elif PAN_PATTERN.search(val_upper):
                        hits.append(
                            f"{cell_ref} — possible PAN ({value.strip()})"
                        )

                    else:
                        key_lower = value.strip().lower()
                        if (
                            COMPANY_NAME_TRIGGER in key_lower
                            and any(t in key_lower for t in COMPANY_NAME_ENTITY_TOKENS)
                        ):
                            # Only flag if adjacent cell (same row, next column) has a value.
                            # Avoids flagging label cells with no corresponding company name.
                            # Scan forward past suppressed merge cells to find
                            # the first non-empty value — the actual company name.
                            next_val = None
                            for lookahead in range(col_idx + 1, len(row)):
                                candidate = row.iloc[lookahead]
                                if (
                                    candidate is not None
                                    and isinstance(candidate, str)
                                    and candidate.strip()
                                    and candidate.strip().lower() not in ("nan", "none", "")
                                ):
                                    next_val = candidate
                                    break
                            if next_val is not None:
                                hits.append(
                                    f"{cell_ref} — company name field ('{value.strip()}')"
                                )

        return hits

    def _substitute_dataframe(self, df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
        """Apply entity name substitution to all string values in a DataFrame."""
        if not mapping:
            return df
        new_df = df.copy()
        for col in new_df.columns:
            new_df[col] = new_df[col].apply(lambda x: self._substitute_value(x, mapping))
        return new_df

    def _substitute_value(self, val, mapping: Dict[str, str]):
        """Substitute entity names in a single cell value. Non-string values pass through."""
        if not isinstance(val, str):
            return val
        result = val
        # Sort by length descending — longer names must replace before shorter substrings fire
        for real_name in sorted(mapping.keys(), key=len, reverse=True):
            result = result.replace(real_name, mapping[real_name])
        return result

    def serialize_value(self, val):
        if isinstance(val, (datetime, date)):
            return val.isoformat()
        return val

    def serialize_key(self, key):
        if isinstance(key, (datetime, date)):
            return key.isoformat()
        return str(key)

    def serialize_records(self, records):
        return [
            {self.serialize_key(k): self.serialize_value(v) for k, v in row.items()}
            for row in records
        ]

    def excel_ingestor(self, state):
        excel_path = state["excel_path"]
        workbook = pd.ExcelFile(excel_path)

        # Step 0 — read all sheets with default header for entity extraction.
        # PII scan has already run at upload time (detect_pii_raw).
        # This read is for entity mapping and substitution only.
        raw_sheets = {}
        for sheet_name in workbook.sheet_names:
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            df = df.dropna(how="all").dropna(axis=1, how="all")
            raw_sheets[sheet_name] = df

        # Build entity mapping. Substitution is applied to all data
        # before Haiku or Sonnet sees anything.
        entity_mapping = self.build_entity_mapping(raw_sheets)

        # Step 1 — sample all sheets for Haiku classification.
        # Read with header=None (preserves raw row order including header row
        # as row 0 — Haiku can see column labels in the preview data).
        # Substitution applied before building samples.
        sheet_samples = {}
        for sheet_name in workbook.sheet_names:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
            df = df.dropna(how="all").dropna(axis=1, how="all")
            df = self._substitute_dataframe(df, entity_mapping)
            sheet_samples[sheet_name] = {
                "columns": [self.serialize_key(col) for col in df.columns],
                "row_count": len(df),
                "preview": self.serialize_records(df.head(15).to_dict(orient="records"))
            }

        # Step 2 — Haiku classifies sheets (sees only substituted data).
        classifications = self.classify_sheets_with_llm(sheet_samples)

        # Step 3 — ingest relevant sheets fully from substituted raw_sheets.
        # No additional disk read — reuse Step 0 data with substitution applied.
        relevant_categories = {
            "financial_statements", "bureau_data",
            "banking", "emi_table",
            "debtor_creditor", "scoring"
        }

        cran_data = {}
        for sheet_name, category in classifications.items():
            if category in relevant_categories:
                df = self._substitute_dataframe(raw_sheets[sheet_name], entity_mapping)
                cran_data[sheet_name] = {
                    "category": category,
                    "data": self.serialize_records(df.to_dict(orient="records"))
                }

        return {"cran_data": cran_data, "entity_mapping": entity_mapping}
