import unittest
import pandas as pd
from data_ingestor import DataIngestor


class TestDetectPIIRaw(unittest.TestCase):
    """Test detect_pii_raw — pre-upload PII scan on raw DataFrames."""

    def setUp(self):
        self.ingestor = DataIngestor()

    def _make_sheets(self, data: dict) -> dict:
        """
        Helper: build raw_sheets dict from {sheet_name: list_of_rows}.
        Each row is a list of cell values (no column headers — header=None style).
        """
        sheets = {}
        for sheet_name, rows in data.items():
            df = pd.DataFrame(rows)
            sheets[sheet_name] = df
        return sheets

    def test_detect_cin_valid_pattern(self):
        """Detect valid CIN pattern in a cell value."""
        sheets = self._make_sheets({
            "Scoring": [[" CIN", "U74999MH2020PLC123456"]]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertTrue(any("CIN" in h for h in hits))
        self.assertTrue(any("Scoring" in h for h in hits))

    def test_detect_cin_cell_reference_format(self):
        """CIN hit includes sheet name and cell reference."""
        sheets = self._make_sheets({
            "Scoring": [["CIN", "U74999MH2020PLC123456"]]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertTrue(any("Scoring!" in h for h in hits))
        # Cell reference format: SheetName!ColRow
        self.assertTrue(any("!" in h for h in hits))

    def test_detect_pan_valid_pattern(self):
        """Detect valid PAN pattern in a cell value."""
        sheets = self._make_sheets({
            "Scoring": [["Borrower PAN", "ABCDE1234F"]]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertTrue(any("PAN" in h for h in hits))

    def test_detect_cin_multiple_sheets(self):
        """Detect CIN across multiple sheets — all reported."""
        sheets = self._make_sheets({
            "Scoring": [["CIN", "U74999MH2020PLC123456"]],
            "Banking": [["Note", "L65990MH2015ABC987654"]],
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        cin_hits = [h for h in hits if "CIN" in h]
        self.assertEqual(len(cin_hits), 2)
        self.assertTrue(any("Scoring" in h for h in cin_hits))
        self.assertTrue(any("Banking" in h for h in cin_hits))

    def test_detect_pan_multiple_sheets(self):
        """Detect PAN across multiple sheets."""
        sheets = self._make_sheets({
            "Scoring": [["PAN", "ABCDE1234F"]],
            "Banking": [["Field", "XYZAB5678G"]],
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        pan_hits = [h for h in hits if "PAN" in h]
        self.assertEqual(len(pan_hits), 2)

    def test_detect_company_name_field(self):
        """Detect company name field label with value."""
        sheets = self._make_sheets({
            "Scoring": [["Company Name", "ABC Industries Ltd"]]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertTrue(any("company name field" in h.lower() for h in hits))

    def test_company_name_aliases(self):
        """Detect all company name field label variants."""
        test_cases = [
            "Company Name",
            "Borrower Name",
            "Entity Name",
            "Business Name",
            "Firm Name",
            "Name of Borrower",
            "Name of the Borrower",
            "Name of Company",
            "Name of the Company",
            "Name of Applicant",
            "Name of the Applicant Unit",
            "Applicant Name",
            "Customer Name",
            "Unit Name",
            "Organisation Name",
        ]
        for field_name in test_cases:
            with self.subTest(field=field_name):
                sheets = self._make_sheets({
                    "Scoring": [[field_name, "Some Company Ltd"]]
                })
                hits = self.ingestor.detect_pii_raw(sheets)
                self.assertTrue(
                    any("company name field" in h.lower() for h in hits),
                    f"Expected company name hit for field '{field_name}'"
                )

    def test_skip_company_name_field_if_empty_value(self):
        """Don't flag company name field if value is empty."""
        sheets = self._make_sheets({
            "Scoring": [["Company Name", ""]]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertEqual(hits, [])

    def test_clean_data_no_hits(self):
        """Return empty list for clean data."""
        sheets = self._make_sheets({
            "Scoring": [["Director", "PERSON_1"], ["Score", "A"]],
            "Financials": [["Revenue", "1000000"], ["Expense", "500000"]],
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertEqual(hits, [])

    def test_non_string_values_skipped(self):
        """Non-string values (numbers, None) are skipped without error."""
        sheets = self._make_sheets({
            "Scoring": [[123456789, None, 95.5]]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertEqual(hits, [])

    def test_all_three_pii_types_in_one_sheet(self):
        """Detect all three PII types when present in same sheet."""
        sheets = self._make_sheets({
            "Scoring": [
                ["CIN", "U74999MH2020PLC123456"],
                ["PAN", "ABCDE1234F"],
                ["Company Name", "ABC Industries Ltd"],
            ]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertTrue(any("CIN" in h for h in hits))
        self.assertTrue(any("PAN" in h for h in hits))
        self.assertTrue(any("company name field" in h.lower() for h in hits))

    def test_empty_sheets(self):
        """Empty sheets dict returns no hits."""
        hits = self.ingestor.detect_pii_raw({})
        self.assertEqual(hits, [])

    def test_cin_case_insensitive(self):
        """CIN detection is case-insensitive."""
        sheets = self._make_sheets({
            "Scoring": [["field", "u74999mh2020plc123456"]]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertTrue(any("CIN" in h for h in hits))

    def test_pan_case_insensitive(self):
        """PAN detection is case-insensitive."""
        sheets = self._make_sheets({
            "Scoring": [["field", "abcde1234f"]]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertTrue(any("PAN" in h for h in hits))

    def test_hit_contains_cell_value(self):
        """Hit message includes the actual cell value for user to locate it."""
        sheets = self._make_sheets({
            "Scoring": [["CIN", "U74999MH2020PLC123456"]]
        })
        hits = self.ingestor.detect_pii_raw(sheets)
        self.assertTrue(any("U74999MH2020PLC123456" in h for h in hits))

    def test_col_to_letter(self):
        """Column index to Excel letter conversion is correct."""
        self.assertEqual(self.ingestor._col_to_letter(0), "A")
        self.assertEqual(self.ingestor._col_to_letter(25), "Z")
        self.assertEqual(self.ingestor._col_to_letter(26), "AA")
        self.assertEqual(self.ingestor._col_to_letter(51), "AZ")
        self.assertEqual(self.ingestor._col_to_letter(52), "BA")


    def test_merged_cell_company_name_detected(self):
        """
        'Name of Borrower' spanning a merged region (A1:B1) with the company name
        in the adjacent cell (C1) must be detected via read_raw_sheets().

        Old pandas read_excel collapsed the merged region to NaN in non-top-left
        cells, so row.iloc[col_idx + 1] was NaN and the adjacent-cell check missed
        the hit.  The openpyxl implementation fills every cell in the merged region
        with the top-left value, making the adjacent cell a non-empty string and
        allowing detect_pii_raw() to fire.
        """
        import tempfile
        import os
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scoring"
        ws["A1"] = "Name of Borrower"
        ws.merge_cells("A1:B1")       # label spans A1:B1; B1 is NaN under pandas
        ws["C1"] = "ABC Industries Ltd"  # value sits one column past the merged region

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            wb.save(tmp_path)
            sheets = self.ingestor.read_raw_sheets(tmp_path)
            hits = self.ingestor.detect_pii_raw(sheets)
            self.assertTrue(
                any("company name field" in h.lower() for h in hits),
                "Expected company name hit from merged-cell layout; got: " + repr(hits),
            )
        finally:
            os.unlink(tmp_path)


if __name__ == "__main__":
    unittest.main()
